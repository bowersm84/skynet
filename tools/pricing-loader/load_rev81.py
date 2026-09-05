#!/usr/bin/env python3
"""
S11 Batch A — Rev 81 price-book loader.
Inputs (all in ./inputs):
  NEW_SALES_TEAM_Skybolt_Pricing_Guide_June_2026.xlsx   the guide (Rev 81)
  2026-09-03_Pricing_Guide_Anomalies.xlsx               Matt's v1 decisions (descriptions, ladders, duplicates)
  2026-09-03_Resale_NoPrice_Proposal.xlsx               Matt's resale decisions
  PartsTable.csv, SaleHistory.csv, CustomerTotals.csv    Fishbowl pull 2026-09-03
  cmp.pkl / tiers.pkl                                    range→SKU resolution, resale set, tier candidates (from the compare round)
Outputs:
  2026-09-05_pricing_seed_rev81.sql   one DO block, dry-run by default
  load_report.md                      counts + every rule applied
Decisions applied: D-PRICE-05..15 (see S11_Implementation_Plan.md §6).
"""
import re, json, pickle, math, datetime as dt
from collections import Counter, defaultdict, OrderedDict
import pandas as pd
from openpyxl import load_workbook

IN = '/home/claude/work/inputs/'
GUIDE = IN + 'NEW_SALES_TEAM_Skybolt_Pricing_Guide_June_2026.xlsx'
OUT_SQL = '/mnt/user-data/outputs/2026-09-05_pricing_seed_rev81.sql'
OUT_REPORT = '/mnt/user-data/outputs/load_report.md'
PREMIER_PCT = 0.97
REV81_LABEL, REV81_DATE = 'Rev 81 — Jun 2026', '2026-05-26'
REV82_LABEL, REV82_DATE, REV82_UPLIFT = 'Rev 82 — Oct 2026', '2026-10-01', 0.15
PINAIR, AIRTRACTOR = 12587, 552          # fb_customer_id (CustomerTotals)
SKIP_PREFIX = ('SK5S5',)
SKIP_PARTS = {'D8-316-709-190'}
RULE_MERGE = {'K': 'D', 'N': 'C', 'P': 'A'}
GENERIC_HEADER = {'skybolt', 'skybolt®', 'part number', 'part#', 'stud part#', 'assembly part#', 'skytanium', 'part number.'}

norm = lambda s: re.sub(r'\s+', '', str(s)).upper()
isnum = lambda v: isinstance(v, (int, float)) and not isinstance(v, bool)
isfx = lambda v: isinstance(v, str) and v.startswith('=')
def q(s):
    if s is None: return 'NULL'
    return "'" + str(s).replace("'", "''") + "'"
def n3(x): return None if x is None else round(float(x), 3)
report = OrderedDict(); counts = Counter()
def note(k, v): report.setdefault(k, []).append(v)

# ---------------------------------------------------------------- guide (formulas + values)
wbf = load_workbook(GUIDE, read_only=True); wbv = load_workbook(GUIDE, read_only=True, data_only=True)
rows_f = list(wbf['Pricing June 2026'].iter_rows(min_row=1, max_row=3568, values_only=True))
rows_v = list(wbv['Pricing June 2026'].iter_rows(min_row=1, max_row=3568, values_only=True))
g = lambda r, i: r[i-1] if len(r) >= i else None

# rule table Q16:W30 with decisions
rules = {}
for r in rows_f[15:30]:
    code = g(r, 17)
    if code: rules[code] = [g(r, c) for c in range(18, 24)]
rules['F'][2] = 0.73;                                   note('rules', 'F Tier 3 filled = 0.73 (matches C)')
for k in ('K', 'N', 'P'): rules.pop(k, None);          note('rules', 'K, N, P retired (K→D, N→C, P→A each_t1_t2)')
rules['KIT'] = [0.80, 0.75, None, None, None, None];    note('rules', 'KIT rule added for the Cloc 2000 Kit (2–4 = ×0.80, 5+ = ×0.75, typed on the sheet)')

# ---------------------------------------------------------------- decisions workbooks
dec = load_workbook(IN + '2026-09-03_Pricing_Guide_Anomalies.xlsx', data_only=True)['Anomalies']
desc_fix = {}
for r in dec.iter_rows(min_row=5, values_only=True):
    if r[0] == 'No description' and r[5] and 'ignore' not in str(r[5]).lower():
        desc_fix[int(r[1])] = str(r[5]).strip()
note('descriptions', f'{len(desc_fix)} descriptions supplied by Matt applied by sheet row')

resale_dec = {}
rp = load_workbook(IN + '2026-09-03_Resale_NoPrice_Proposal.xlsx', data_only=True).active
for r in rp.iter_rows(min_row=5, values_only=True):
    if r[0]: resale_dec[norm(r[0])] = str(r[10]).strip().strip('"') if r[10] is not None else ''

cmp = pickle.load(open(IN + 'cmp.pkl', 'rb')); tiers = pickle.load(open(IN + 'tiers.pkl', 'rb'))
prod = pd.read_csv(IN + 'PartsTable.csv', encoding='cp1252'); prod = prod.loc[:, ~prod.columns.str.startswith('Unnamed')]
prod['key'] = prod.product_num.map(norm); fbid = dict(zip(prod.key, prod.product_id)); fbprice = dict(zip(prod.key, prod.list_price))
fb_keys = set(prod.key)

# ---------------------------------------------------------------- walk the sheet: sections, ladders, items
def ladder_from_header(r):
    """I..K qty columns (numeric label → min qty); L..N tier columns (any label → tier1..3). Returns (code, columns)."""
    cols = []
    for c, key in zip((9, 10, 11), ('a', 'b', 'c')):
        v = g(r, c)
        m = re.match(r'^\s*(\d+)', str(v)) if v is not None and not isfx(v) else None
        if m: cols.append({'kind': 'qty', 'min': int(m.group(1))})
    for c, key in zip((12, 13, 14), ('tier1', 'tier2', 'tier3')):
        v = g(r, c)
        if v is not None and str(v).strip() and not isfx(v): cols.append({'kind': 'tier', 'key': key})
    # formula cells in L..N under a numeric-only header: the rows still use the tier multipliers → keep tier columns
    if not any(c['kind'] == 'tier' for c in cols) and any(isfx(g(r, c)) for c in (12, 13, 14)):
        for key in ('tier1', 'tier2', 'tier3'): cols.append({'kind': 'tier', 'key': key})
    out = []
    for c in cols:
        if c['kind'] == 'qty': out.append({'key': f"q{c['min']}", 'kind': 'qty', 'min': c['min'], 'label': str(c['min'])})
        else: out.append({'key': c['key'], 'kind': 'tier', 'label': {'tier1': 'Tier 1', 'tier2': 'Tier 2', 'tier3': 'Tier 3'}[c['key']]})
    sig = '_'.join(x['key'] for x in out)
    code = 'standard' if sig == 'q100_q300_q500_tier1_tier2_tier3' else (sig or 'none')
    return code, out

ladders = {'standard': [{'key': 'q100', 'kind': 'qty', 'min': 100, 'label': '100'}, {'key': 'q300', 'kind': 'qty', 'min': 300, 'label': '300'},
                        {'key': 'q500', 'kind': 'qty', 'min': 500, 'label': '500'}, {'key': 'tier1', 'kind': 'tier', 'label': 'Tier 1'},
                        {'key': 'tier2', 'kind': 'tier', 'label': 'Tier 2'}, {'key': 'tier3', 'kind': 'tier', 'label': 'Tier 3'}],
           'each_t1_t2': [{'key': 'tier1', 'kind': 'tier', 'label': 'Tier 1'}, {'key': 'tier2', 'kind': 'tier', 'label': 'Tier 2'}],
           'none': [],
           'kit_2_5': [{'key': 'q2', 'kind': 'qty', 'min': 2, 'label': '2–4'}, {'key': 'q5', 'kind': 'qty', 'min': 5, 'label': '5+'}]}
sections = []            # dicts: name, sort, row, kind
items = []               # dicts
title_pending = None; cur_section = None; cur_ladder = 'standard'
PN_LIKE = re.compile(r'^[A-Z0-9][A-Z0-9\-\./]{3,}$')
TITLE_WORDS = re.compile(r'(Series|Skybolt|Cloc|CLoc|ZLoc|Buttons|Ejector|Grommet|Retainer|Tools|Receiver|Receptacle|Kit|Sets|Nutplate|Steel|Stainless|Skytanium|Discontinued|Head|Stud|Insert|Washer|Adhesive|Doubler|Tabs|#\d)', re.I)
for n, (rf, rv) in enumerate(zip(rows_f, rows_v), 1):
    B = g(rf, 2); H = g(rf, 8)
    Bs = str(B).strip() if isinstance(B, str) else ''
    if Bs.startswith(SKIP_PREFIX) or Bs in SKIP_PARTS: counts['skipped_sk5s5'] += 1; continue
    if H == 'Each':                                   # header row → new section + ladder
        code, cols = ladder_from_header(rf)
        if code not in ladders: ladders[code] = cols
        own = Bs if Bs and Bs.lower().rstrip('.') not in GENERIC_HEADER and not PN_LIKE.match(Bs) else None
        name = own or title_pending
        if name is None and cur_section is not None:
            if code == cur_ladder: continue                       # continuation header inside the same section
            name = cur_section['name']; counts['continuation_sections'] += 1
        cur_section = {'name': re.sub(r'\s{2,}', ' ', name or 'Untitled'), 'sort': len(sections) + 1, 'row': n, 'kind': 'catalog'}
        sections.append(cur_section); cur_ladder = code; title_pending = None
        continue
    priced = isnum(H) or isfx(H)
    if Bs and not priced and H is None:
        if PN_LIKE.match(Bs):                          # unpriced part → item, no_price (D-PRICE-12)
            if cur_section is None: continue
            items.append(dict(row=n, pn=Bs, desc=(desc_fix.get(n) or (str(g(rf, 7)).strip() if g(rf, 7) else None)), each=None, rule=None, ladder=cur_ladder, section=cur_section, status='no_price',
                              dfar=(g(rf, 1) == 'Y'), xa=g(rf, 3), xl=g(rf, 4), nsn=g(rf, 5), cess=g(rf, 6), premier=False, range_of=None, kit=None))
            counts['unpriced_parts'] += 1
        elif TITLE_WORDS.search(Bs) or len(Bs) > 24:
            title_pending = re.sub(r'\s{2,}', ' ', Bs)
        continue
    if not (Bs and priced): continue
    if cur_section is None: continue
    each = g(rv, 8) if isfx(H) else H
    if isfx(H): note('each_formula', f'row {n} {Bs}: Each formula evaluated to {each}')
    if not isnum(each) or each <= 0:
        counts['dropped_zero_each'] += 1; note('dropped', f'row {n} {Bs}: Each={each} dropped'); continue
    rule = g(rf, 25); ladder = cur_ladder; status = 'priced'; kit = None
    colO = g(rf, 15); premier = isnum(colO) or isfx(colO)
    I = g(rf, 9)
    if rule in (None, ''):
        if Bs.lower().startswith('cloc 2000 kit'):
            rule = 'KIT'; ladder = 'kit_2_5'
        elif isfx(I) and '+' in I and '*' not in I:    # SET row: =I64+I657+I744
            refs = [int(x) for x in re.findall(r'[A-Z]+(\d+)', I)]
            comps = []
            for rr in refs:
                cb = g(rows_f[rr-1], 2); comps.append(str(cb).strip())
            kit = comps; status = 'component_sum'; rule = None
        else:
            counts['dropped_no_rule'] += 1; note('dropped', f'row {n} {Bs}: no rule and not a set'); continue
    else:
        if rule == 'P': ladder = 'each_t1_t2'
        rule = RULE_MERGE.get(rule, rule)
        if rule not in rules: note('dropped', f'row {n} {Bs}: unknown rule {rule}'); continue
    desc = desc_fix.get(n) or (str(g(rf, 7)).strip() if g(rf, 7) else None)
    items.append(dict(row=n, pn=Bs, desc=desc, each=n3(each), rule=rule, ladder=ladder, section=cur_section, status=status,
                      dfar=(g(rf, 1) == 'Y'), xa=g(rf, 3), xl=g(rf, 4), nsn=g(rf, 5), cess=g(rf, 6), premier=premier, range_of=None, kit=kit))
counts['guide_rows_read'] = len(items)

# ---------------------------------------------------------------- ranges → SKUs
range_hits = {r[0]: r for r in cmp['range_map']}
def parse_range(pn):
    s = pn.replace('_', '-').strip(); parts = re.split(r'\s*to\s+', s, maxsplit=1, flags=re.I)
    if len(parts) != 2: return None
    left, right = parts[0].strip().rstrip('-'), parts[1].strip().lstrip('-').strip()
    lm, rm = list(re.finditer(r'\d+', left)), list(re.finditer(r'\d+', right))
    if not lm or not rm: return None
    l, r = lm[-1], rm[-1]
    pre = re.sub(r'\s+', '', left[:l.start()].rstrip('-')).upper(); suf = (right[r.end():] or left[l.end():]).upper()
    return pre, int(l.group()), int(r.group()), suf
def resolve(pn):
    pa = parse_range(pn)
    if not pa: return []
    pre, a, b, suf = pa; hits = []
    for k in fb_keys:
        if not k.startswith(pre): continue
        m = re.match(r'^(\d+)([A-Z0-9]*)$', k[len(pre):].lstrip('-'))
        if m and a <= int(m.group(1)) <= b and m.group(2) == suf: hits.append(k)
    return sorted(hits, key=lambda x: int(re.search(r'(\d+)[A-Z0-9]*$', x).group(1)))
expanded = []
for it in items:
    if re.search(r'(\s|\d)to\s', it['pn'], re.I):
        if it['row'] in range_hits and range_hits[it['row']][3]: hits = range_hits[it['row']][3].split(', ')
        else: hits = resolve(it['pn'])
        if not hits: counts['ranges_no_product'] += 1; note('ranges_dropped', f"row {it['row']} {it['pn']}: no Fishbowl product"); continue
        for h in hits:
            pnum = prod.loc[prod.key == h, 'product_num'].iloc[0]
            e = dict(it); e['pn'] = pnum; e['range_of'] = it['pn']; expanded.append(e)
        counts['range_rows'] += 1; counts['range_skus'] += len(hits)
    else: expanded.append(it)
items = expanded

# ---------------------------------------------------------------- Matt's removals (resale sheet) apply everywhere
removed = {k for k, d in resale_dec.items() if d.lower() == 'remove'}
before = len(items); items = [it for it in items if norm(it['pn']) not in removed]
if before - len(items): note('removed', f"{before-len(items)} guide rows removed per the resale decisions: " + ', '.join(sorted(removed)))
# ---------------------------------------------------------------- duplicates: higher Each wins
by = OrderedDict()
for it in items:
    k = norm(it['pn'])
    if k in by:
        old = by[k]
        if (it['each'] or 0) > (old['each'] or 0): by[k] = it
        counts['duplicates_resolved'] += 1; note('duplicates', f"{it['pn']}: rows {old['row']} & {it['row']} → kept row {by[k]['row']} (Each {by[k]['each']})")
    else: by[k] = it
items = list(by.values())

# ---------------------------------------------------------------- resale section
resale_sec = {'name': 'Resale Items', 'sort': len(sections) + 1, 'row': None, 'kind': 'resale'}; sections.append(resale_sec)
ns = cmp['notsheet']; res = ns[ns.active & ns.sold]
alias = {}   # resale part → guide part whose pricing it adopts
for x in res.itertuples():
    k = norm(x.product_num); d = resale_dec.get(k, '')
    if d.lower() == 'remove': counts['resale_removed'] += 1; continue
    if k in by: continue                                     # already on the guide (range expansion)
    price = n3(x.list_price) if pd.notna(x.list_price) and float(x.list_price) > 0 else None
    status = 'priced'
    if price is None:
        try: price = n3(float(d)); note('resale_priced_by_matt', f'{x.product_num} = {price}')
        except ValueError:
            m = re.search(r'(SK[A-Z0-9\-]+)', d)
            if d.startswith('See pricing on row'): alias[k] = norm('SK212-12SD')
            elif m and norm(m.group(1)) in by: alias[k] = norm(m.group(1))
            elif 'guide' in d.lower(): alias[k] = norm('SK203A01AE')
            elif 'Adopt' in d: alias[k] = norm('SK212-12S')
            if k in alias:
                src = by[alias[k]]; it = dict(src); it.update(pn=x.product_num, row=None, range_of=None, notes=f"priced as {src['pn']} per Matt 2026-09-03")
                items.append(it); note('resale_aliased', f"{x.product_num} → {src['pn']}"); continue
            status = 'no_price'; counts['resale_no_price'] += 1
    items.append(dict(row=None, pn=x.product_num, desc=(x.product_desc if isinstance(x.product_desc, str) else None), each=price, rule=None, ladder='none',
                      section=resale_sec, status=status, dfar=False, xa=None, xl=None, nsn=None, cess=None, premier=False, range_of=None, kit=None))
    counts['resale_items'] += 1

# ---------------------------------------------------------------- premier snap / exceptions
exc = []
for it in items:
    if it['status'] != 'priced' or not it['premier']: continue
    t3m = rules[it['rule']][5]
    if it['pn'].startswith('SK201-') or it['pn'].startswith('SK40R17'):
        rf = rows_f[it['row']-1]; cur = g(rf, 15)
        pct = round(float(cur) / (it['each'] * t3m), 4)
        cust = PINAIR if it['pn'].startswith('SK201-') else AIRTRACTOR
        exc.append((cust, it['pn'], pct, f"special Premier kept from Rev 81 ({cur:.3f} vs T3 {it['each']*t3m:.3f})"))
counts['premier_parts'] = sum(1 for it in items if it['premier']); counts['exceptions'] = len(exc)

# ---------------------------------------------------------------- tiers
tier_rows = []
for x in tiers['cust'].itertuples():
    m = re.search(r'Tier (\d)', str(x.inferred))
    if m: tier_rows.append((int(x.fb_customer_id), f'tier{m.group(1)}', f"seeded from paid-price analysis 2026-09-03 ({x.inferred}, {x.pct_tier}% of revenue at tier) — confirm"))
tier_rows.append((PINAIR, 'premier', 'seeded 2026-09-03: negotiated below Tier 3 on SK201 receptacles — confirm'))
tier_rows.append((AIRTRACTOR, 'premier', 'seeded 2026-09-03: negotiated below Tier 3 on SK40R17 — confirm'))
counts['tier_seeds'] = len(tier_rows)

used_l = {it['ladder'] for it in items} | {'standard','each_t1_t2','none','kit_2_5'}
ladders = {k: v for k, v in ladders.items() if k in used_l}
# ---------------------------------------------------------------- write SQL
L = []
L.append(f"""/* ============================================================================
   S11 Batch A — Rev 81 price-book SEED. Generated {dt.date.today()} by tools/pricing-loader/load_rev81.py
   Supabase SQL Editor (TEST first). Requires 2026-09-04_pricing_schema.sql.
   DRY RUN BY DEFAULT: the block performs every insert, prints the counts, then rolls itself back.
   Set  v_dry_run := false  and re-run to commit. Guarded: refuses to run twice.
   Creates: book '{REV81_LABEL}' (active, effective {REV81_DATE}), '{REV82_LABEL}' (scheduled {REV82_DATE}, catalog Each x{1+REV82_UPLIFT}),
            rules, ladders, {len(sections)} sections, {len(items)} items, kit components, {len(exc)} exceptions, {len(tier_rows)} tier seeds.
   ============================================================================ */
DO $$
DECLARE
  v_dry_run boolean := true;
  v_b81 uuid; v_b82 uuid; v_sec uuid; v_item uuid; n_items int; n_kits int; n_sec int;
  v_secs jsonb := '{{}}'::jsonb;
BEGIN
  IF EXISTS (SELECT 1 FROM public.price_books WHERE rev_label = {q(REV81_LABEL)}) THEN
    RAISE EXCEPTION 'GUARD: book % already exists — seed already applied', {q(REV81_LABEL)};
  END IF;
  INSERT INTO public.price_books (rev_label, effective_from, status, source, premier_pct, notes)
  VALUES ({q(REV81_LABEL)}, {q(REV81_DATE)}, 'active', 'guide_rev81', {PREMIER_PCT}, 'Seeded from NEW_SALES_TEAM_Skybolt_Pricing_Guide_June_2026.xlsx (Rev 81 - MB) with the 2026-09-03 anomaly decisions')
  RETURNING id INTO v_b81;
""")
L.append("  /* rules (D-PRICE-07) */\n")
for code, v in rules.items():
    vals = ', '.join('NULL' if x in (None, '') else str(x) for x in v)
    L.append(f"  INSERT INTO public.price_rules VALUES (v_b81, {q(code)}, {vals}, NULL);\n")
L.append("  /* ladders (D-PRICE-08) */\n")
for code, cols in ladders.items():
    L.append(f"  INSERT INTO public.price_ladders VALUES (v_b81, {q(code)}, {q(code)}, {q(json.dumps(cols))}::jsonb);\n")
L.append("  /* sections */\n")
for s in sections:
    L.append(f"  INSERT INTO public.price_sections (book_id, name, sort, kind, source_row) VALUES (v_b81, {q(s['name'])}, {s['sort']}, {q(s['kind'])}, {s['row'] if s['row'] else 'NULL'}) RETURNING id INTO v_sec; v_secs := v_secs || jsonb_build_object({q(str(s['sort']))}, v_sec);\n")
L.append("  /* items */\n  INSERT INTO public.price_items (book_id, section_id, part_number, fb_product_id, description, list_price, rule_code, ladder_code, has_premier, dfar, xref_arconic, xref_lisi, nsn, cessna, sort, status, source_row, range_of, notes) VALUES\n")
vals = []
for i, it in enumerate(items, 1):
    k = norm(it['pn']); fb = fbid.get(k)
    vals.append(f"  (v_b81, (v_secs->>{q(str(it['section']['sort']))})::uuid, {q(it['pn'])}, {int(fb) if fb is not None and not (isinstance(fb,float) and math.isnan(fb)) else 'NULL'}, {q(it['desc'])}, "
                f"{it['each'] if it['each'] is not None else 'NULL'}, {q(it['rule'])}, {q(it['ladder'])}, {str(bool(it['premier'])).lower()}, {str(bool(it['dfar'])).lower()}, "
                f"{q(it['xa'])}, {q(it['xl'])}, {q(it['nsn'])}, {q(it['cess'])}, {i}, {q(it['status'])}, {it['row'] if it['row'] else 'NULL'}, {q(it['range_of'])}, {q(it.get('notes'))})")
L.append(',\n'.join(vals) + ';\n')
L.append("  /* kit components (D-PRICE-09) */\n")
for it in items:
    if it['kit']:
        for comp in it['kit']:
            L.append(f"  INSERT INTO public.kit_components (item_id, component_part_number, qty) SELECT id, {q(comp)}, 1 FROM public.price_items WHERE book_id = v_b81 AND part_key = {q(norm(it['pn']))};\n")
L.append("  /* link SkyNet parts / kit SKUs by part number */\n"
         "  UPDATE public.price_items i SET part_id = p.id FROM public.parts p WHERE i.book_id = v_b81 AND i.part_id IS NULL AND upper(regexp_replace(p.part_number, '\\s', '', 'g')) = i.part_key;\n"
         "  UPDATE public.price_items i SET kit_sku_id = k.id FROM public.kit_skus k WHERE i.book_id = v_b81 AND i.kit_sku_id IS NULL AND upper(regexp_replace(k.part_number, '\\s', '', 'g')) = i.part_key;\n")
L.append("  /* customer x part exceptions (D-PRICE-06) */\n")
for cust, pn, pct, notetxt in exc:
    L.append(f"  INSERT INTO public.price_exceptions (fb_customer_id, part_number, mode, value, note, effective_from) VALUES ({cust}, {q(pn)}, 'pct_of_tier3', {pct}, {q(notetxt)}, {q(REV81_DATE)});\n")
L.append("  /* tier seeds — candidates, to be confirmed in the portal */\n")
for cust, tier, notetxt in tier_rows:
    L.append(f"  INSERT INTO public.customer_pricing (fb_customer_id, tier, effective_from, note) VALUES ({cust}, {q(tier)}, {q(REV81_DATE)}, {q(notetxt)});\n")
L.append(f"""  /* Rev 82 — the Oct 1 book (D-PRICE-15) */
  v_b82 := public.pricing_clone_book(v_b81, {q(REV82_LABEL)}, {q(REV82_DATE)}, {REV82_UPLIFT}, 'Oct 1 2026 increase: every catalog Each x1.15; resale unchanged');
  PERFORM public.pricing_publish_book(v_b82, {q(REV82_DATE)});

  SELECT count(*) INTO n_items FROM public.price_items WHERE book_id = v_b81;
  SELECT count(*) INTO n_kits  FROM public.kit_components kc JOIN public.price_items i ON i.id = kc.item_id WHERE i.book_id = v_b81;
  SELECT count(*) INTO n_sec   FROM public.price_sections WHERE book_id = v_b81;
  RAISE NOTICE 'Rev 81 % / Rev 82 %: sections=% items=% kit_components=% expected items={len(items)} kit_components={sum(len(it['kit']) for it in items if it['kit'])}', v_b81, v_b82, n_sec, n_items, n_kits;
  IF v_dry_run THEN
    RAISE EXCEPTION 'DRY RUN OK — sections=% items=% kit_components=% (rolled back; set v_dry_run := false to apply)', n_sec, n_items, n_kits;
  END IF;
END $$;
""")
open(OUT_SQL, 'w', encoding='utf-8').write(''.join(L))

# ---------------------------------------------------------------- report
r = [f"# Rev 81 load report — {dt.date.today()}\n", "## Counts\n"]
for k, v in sorted(counts.items()): r.append(f"- {k}: {v}")
r.append(f"- sections: {len(sections)}  items: {len(items)}  ladders: {len(ladders)}  rules: {len(rules)}")
status = Counter(it['status'] for it in items); r.append(f"- item status: {dict(status)}")
r.append(f"- items per section kind: {dict(Counter(it['section']['kind'] for it in items))}")
r.append("\n## Ladders derived from section headers\n")
for code, cols in ladders.items(): r.append(f"- `{code}`: " + ' | '.join(c['label'] for c in cols) + f"  — used by {sum(1 for it in items if it['ladder']==code)} items")
r.append("\n## Rules\n")
for code, v in rules.items(): r.append(f"- {code}: {v}  — {sum(1 for it in items if it['rule']==code)} items")
r.append("\n## Exceptions (pct of Tier 3)\n")
for c, pn, pct, nt in exc: r.append(f"- customer {c} × {pn}: {pct}  ({nt})")
for k, v in report.items():
    r.append(f"\n## {k}\n"); r.extend(f"- {x}" for x in v[:200])
    if len(v) > 200: r.append(f"- … {len(v)-200} more")
open(OUT_REPORT, 'w', encoding='utf-8').write('\n'.join(r))
print(json.dumps(counts, indent=1)); print('sections', len(sections), 'items', len(items), 'ladders', list(ladders))
